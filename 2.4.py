Главный бухгалтер компании "Рога и копыта" случайно удалил ведомость с начисленной зарплатой. К счастью, у него сохранились расчётные листки всех сотрудников. Помогите по этим расчётным листкам восстановить зарплатную ведомость. Архив с расчётными листками доступен по ссылке https://stepik.org/media/attachments/lesson/245299/rogaikopyta.zip (вы можете скачать и распаковать его вручную или самостоятельно научиться делать это с помощью скрипта на Питоне).

Ведомость должна содержать 1000 строк, в каждой строке должно быть указано ФИО сотрудника и, через пробел, его зарплата. Сотрудники должны быть упорядочены по алфавиту.

from io import BytesIO
import openpyxl
import zipfile
from urllib.request import urlopen

resp = urlopen("https://stepik.org/media/attachments/lesson/245299/rogaikopyta.zip")
employees = {}
with zipfile.ZipFile(BytesIO(resp.read())) as zipfile:
    for file in zipfile.namelist():
        with zipfile.open(file) as book:
            wb = openpyxl.load_workbook(book)
            ws = wb.worksheets[0]
            employees[ws[2][1].value] = employees.get(ws[2][1].value, 0) + int(ws[2][3].value)
for employee in sorted(employees.keys()):
    print(employee + ' ' + str(employees[employee]))
    
# вывод ведомости в .xlsx файл:
#bill = openpyxl.Workbook()
#ws = bill.active
#ws.cell(column=1, row=1).value = 'ФИО'
#ws.cell(column=2, row=1).value = 'Начислено'
#for i, employee in enumerate(sorted(employees.keys()), start=2):
    #ws.cell(column=1, row=i).value = employee
    #ws.cell(column=2, row=i).value = employees[employee]
#bill.save('stepik-course4519-2.4-bill.xlsx')
